import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import time
from datetime import date
import re


header_num_dic = {"User Name":0, "Job Role": 1, "Report into": 2, "Company": 3,
                  "Timezone": 4, "Company Size": 5, "Team Size": 6, "Start Time": 7,
                  "Last Sync":8, "Meeting Title":9, "Number of Guests": 10, "Nudge Requested Date": 11,
                  "Goals": 12, "Framework": 13, "Nudge ID": 14, "Relevant": 15}


def extract_from_excel(filename, variable_list):
    wb = load_workbook(filename)
    ws = wb["Master"]
    ws_by_rows_gen = ws.iter_rows(values_only = True)
    meeting_list = []
    meeting_info = next(ws_by_rows_gen)
    for i in range(1, ws.max_row):
        meeting_info = next(ws_by_rows_gen)
        meeting_dic = {}
        for variable in variable_list:
            if variable in header_num_dic:
                meeting_dic[variable] = meeting_info[header_num_dic[variable]]
        meeting_dic = dates_and_times_corrector(meeting_dic)
        meeting_list += [meeting_dic]
    return meeting_list


#If you ask for start time you also need to ask for timezone
def dates_and_times_corrector(meeting_dic):
    if "Start Time" in meeting_dic:
        start_datetime = start_datetime_corrector(meeting_dic["Start Time"])
        if meeting_dic["Timezone"][1] == "+":
           time_change = int(re.findall(r"\d+", meeting_dic["Timezone"])[0])
        elif meeting_dic["Timezone"][1] == "-":
            time_change = -int(re.findall(r"\d+", meeting_dic["Timezone"])[0])
        else:
            time_change = 0
        start_datetime = start_datetime + datetime.timedelta(hours = time_change)
        meeting_dic["Start Time"] = start_datetime
    return meeting_dic
                                                                
def start_datetime_corrector(start_string):
    start_year = int(start_string[0:4])
    start_month = int(start_string[5:7])
    start_day = int(start_string[8:10])
    start_hour = int(start_string[11:13])
    start_minute = int(start_string[14:16])
    return datetime.datetime(start_year, start_month, start_day,start_hour,start_minute)



def split_by_user(meeting_list, user_name):
    return [meeting_dic for meeting_dic in meeting_list if meeting_dic["User Name"] == user_name]

def write_diary(users_meetings, worksheet):
    meetings_by_week = seperate_by_week(users_meetings)
    start_week_num = meetings_by_week[0][0]["Start Time"].isocalendar()[1]
    week_date = date.fromisocalendar(2020, start_week_num, 1)
    for week in meetings_by_week:
        week_date = week_date + datetime.timedelta(7)
        write_week(week, worksheet, week_date)
    return

def write_week(week, worksheet, start_date):
    start_row = worksheet.max_row
    if start_row != 1:
        start_row += 2
    col_num = 1
    meetings_by_day = seperate_by_day(week)
    for index in range(0,7):
        write_day(meetings_by_day[index], start_row, col_num, worksheet, start_date + datetime.timedelta(index))
        col_num += 4
    return

def write_day(day, start_row, col_num, worksheet, day_date):
    row_num = start_row
    day.sort(key = lambda x: x["Start Time"])
    a = worksheet.cell(row_num, col_num, value = day_date)
    a = worksheet.cell(row_num, col_num+1, value = days_list[day_date.isocalendar()[2]-1])
    
    for meeting in day:
        row_num += 1
        a = worksheet.cell(row_num, col_num, meeting["Start Time"].strftime("%H:%M"))
        a = worksheet.cell(row_num, col_num+1, meeting["Meeting Title"])
        a = worksheet.cell(row_num, col_num+2, meeting["Number of Guests"])
    return


#Takes a users' list of meetings and seperates them by week
def seperate_by_week(users_meetings):
    meetings_by_week = []
    week_nums_list = range(min([elem["Start Time"].isocalendar()[1] for elem in users_meetings]), max([elem["Start Time"].isocalendar()[1] for elem in users_meetings])+1)
    for week_num in week_nums_list:
        meetings_by_week += [ [elem for elem in users_meetings if elem["Start Time"].isocalendar()[1] == week_num] ]
    return meetings_by_week

#Takes a week's worth of meetings and seperates them by day
def seperate_by_day(week_list):
    meetings_by_day = []
    for day_num in range(1,8):
        meetings_by_day += [ [elem for elem in week_list if elem["Start Time"].isocalendar()[2] == day_num] ]
    return meetings_by_day


def write_cal_all(filename):
    wb = load_workbook(filename)
    meeting_list = extract_from_excel("Multiplii Data 15072020.xlsx", ["User Name", "Timezone", "Start Time", "Meeting Title", "Number of Guests"])
    for user in list(set([elem["User Name"] for elem in meeting_list])):
        user_meetings = split_by_user(meeting_list, user)
        ws = wb.create_sheet(user)
        write_diary(user_meetings, ws)
    wb.save(filename)



days_list = ["Mon", "Tue", "Wed", "Thur", "Fri", "Sat", "Sun"]

